import pytest
from pathlib import Path
from openpyxl import load_workbook
from models.finding import Finding
from synthesizer.scorecard import write_scorecard, write_gap_register

FINDINGS = [
    Finding(framework="ISO 27001", control_id="A.8.8", control_name="Patch mgmt",
            status="Gap", severity="High", evidence="Monthly cycle.", recommendation="Add critical SLA."),
    Finding(framework="DORA", control_id="Art.9", control_name="ICT risk",
            status="Compliant", severity="Info", evidence="Framework documented.", recommendation="No action."),
]


def test_write_scorecard_creates_file(tmp_path):
    out = tmp_path / "scorecard.xlsx"
    scorecard = {"ISO 27001": {"rag": "Red", "total": 1, "gaps": 1, "critical": 0, "high": 1, "partial": 0},
                 "DORA": {"rag": "Green", "total": 1, "gaps": 0, "critical": 0, "high": 0, "partial": 0}}
    write_scorecard(scorecard, out)
    assert out.exists()


def test_write_scorecard_has_rag_values(tmp_path):
    out = tmp_path / "scorecard.xlsx"
    scorecard = {"ISO 27001": {"rag": "Red", "total": 2, "gaps": 1, "critical": 0, "high": 1, "partial": 0}}
    write_scorecard(scorecard, out)
    wb = load_workbook(out)
    ws = wb.active
    values = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)]
    assert "Red" in values


def test_write_gap_register_creates_file(tmp_path):
    out = tmp_path / "gaps.xlsx"
    write_gap_register(FINDINGS, out)
    assert out.exists()


def test_write_gap_register_has_findings(tmp_path):
    out = tmp_path / "gaps.xlsx"
    write_gap_register(FINDINGS, out)
    wb = load_workbook(out)
    ws = wb.active
    all_values = [ws.cell(row=r, column=c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    assert "ISO 27001" in all_values
