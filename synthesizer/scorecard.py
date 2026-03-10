from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from models.finding import Finding

RAG_FILLS = {
    "Red":   PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "Amber": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "Green": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}

SEVERITY_FILLS = {
    "Critical": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "High":     PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
    "Medium":   PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "Low":      PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
    "Info":     PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")


def _style_header(ws, row: int, columns: list[str]):
    for col, value in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_scorecard(scorecard: dict, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Scorecard"

    headers = ["Framework", "RAG Status", "Total Controls", "Gaps", "Critical", "High", "Partial"]
    _style_header(ws, 1, headers)

    for row, (framework, data) in enumerate(scorecard.items(), start=2):
        ws.cell(row=row, column=1, value=framework)
        rag_cell = ws.cell(row=row, column=2, value=data["rag"])
        rag_cell.fill = RAG_FILLS.get(data["rag"], PatternFill())
        ws.cell(row=row, column=3, value=data.get("total", 0))
        ws.cell(row=row, column=4, value=data.get("gaps", 0))
        ws.cell(row=row, column=5, value=data.get("critical", 0))
        ws.cell(row=row, column=6, value=data.get("high", 0))
        ws.cell(row=row, column=7, value=data.get("partial", 0))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    wb.save(output_path)


def write_gap_register(findings: list[Finding], output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Register"

    headers = ["Framework", "Control ID", "Control Name", "Status", "Severity", "Evidence", "Recommendation"]
    _style_header(ws, 1, headers)

    for row, f in enumerate(findings, start=2):
        ws.cell(row=row, column=1, value=f.framework)
        ws.cell(row=row, column=2, value=f.control_id)
        ws.cell(row=row, column=3, value=f.control_name)
        ws.cell(row=row, column=4, value=f.status)
        sev_cell = ws.cell(row=row, column=5, value=f.severity)
        sev_cell.fill = SEVERITY_FILLS.get(f.severity, PatternFill())
        ws.cell(row=row, column=6, value=f.evidence)
        ws.cell(row=row, column=7, value=f.recommendation)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 30

    wb.save(output_path)
