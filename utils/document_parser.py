from pathlib import Path


class DocumentParseError(Exception):
    pass


# Hard bound on any single input file. A 50 MB cap is generous for vendor
# questionnaires / SOC 2 reports / ISO certs (typical real-world docs are
# well under 10 MB) while keeping a crafted PDF or zip-bomb XLSX from
# driving the host (or the 2-GiB-RAM Firecracker VM) into OOM before
# parsing even begins.
MAX_FILE_BYTES = 50 * 1024 * 1024

# PDFs with unusually high page counts are almost always either (a) an
# accidentally-included full regulatory text or (b) a crafted input
# aimed at pypdf's per-page allocation. Cap what we'll process — docs
# exceeding this are truncated rather than parsed in full.
MAX_PDF_PAGES = 2000

# XLSX with massive sheet counts is another decompression-blowup vector.
MAX_XLSX_SHEETS = 50


def _enforce_size_cap(path: Path) -> None:
    size = path.stat().st_size
    if size > MAX_FILE_BYTES:
        raise DocumentParseError(
            f"file exceeds {MAX_FILE_BYTES // (1024 * 1024)} MB size cap: "
            f"{path.name} is {size // (1024 * 1024)} MB"
        )


def parse_document(path: Path) -> str:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    _enforce_size_cap(path)

    suffix = path.suffix.lower()

    if suffix == ".txt":
        return path.read_text(encoding="utf-8", errors="replace")

    if suffix == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        pages = reader.pages
        if len(pages) > MAX_PDF_PAGES:
            # Truncate rather than reject — a legitimate long doc is a
            # more likely explanation than a bomb for edge cases, and
            # the caller gets a usable subset. Log via the returned
            # text so the auditor sees what happened.
            truncated_text = "\n".join(
                pages[i].extract_text() or "" for i in range(MAX_PDF_PAGES)
            )
            return (
                truncated_text
                + f"\n\n[document_parser: truncated — source has {len(pages)} "
                f"pages, processed first {MAX_PDF_PAGES}]"
            )
        return "\n".join(page.extract_text() or "" for page in pages)

    if suffix == ".docx":
        from docx import Document
        doc = Document(str(path))
        return "\n".join(para.text for para in doc.paragraphs)

    if suffix in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(str(path), data_only=True)
        if len(wb.worksheets) > MAX_XLSX_SHEETS:
            raise DocumentParseError(
                f"workbook has {len(wb.worksheets)} sheets — exceeds cap of "
                f"{MAX_XLSX_SHEETS}; refusing to parse {path.name}"
            )
        lines = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                line = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if line.strip():
                    lines.append(line)
        return "\n".join(lines)

    raise DocumentParseError(f"Unsupported file format: {suffix}")


def parse_documents(paths: list[Path]) -> str:
    parts = []
    for path in paths:
        path = Path(path)
        text = parse_document(path)
        parts.append(f"=== FILE: {path.name} ===\n{text}")
    return "\n\n".join(parts)
